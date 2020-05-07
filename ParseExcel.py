#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May  5 10:36:32 2020

This script search and remove non-printable characters and trailing whitespaces

@author: pinin
"""
import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

class Parse_Excel():
    def __init__(self):
        root = tk.Tk()
        root.title("Parse Excel File")

        frame0 = tk.Frame(root)
        frame0.pack()

        # Create buttons
        browseButton = tk.Button(frame0, text="Import Excel File",
                                 command=self.getExcel)
        browseButton.pack(side = tk.LEFT)
        
        parseButton = tk.Button(frame0, text = "Find Whitespace",
                                command = self.getNonprintable)
        parseButton.pack(side = tk.LEFT)
 
        saveButton = tk.Button(frame0, text = "Remove and Save",
                                command = self.removeNonprintable)
        saveButton.pack(side = tk.LEFT)
        
        # Display file path
        frame1 = tk.Frame(root)
        frame1.pack()
        
        self.fileLabel = tk.StringVar()
        self.fileLabel.set("Open an Excel file (.xlsx)")
        tk.Label(frame1, textvariable = self.fileLabel).pack()
        
        self.filePath = "" 
        
        # Display found result
        frame2 = tk.Frame(root)
        frame2.pack()
        
        self.text = tk.StringVar()       
        message = tk.Message(frame2, textvariable = self.text)
        message.pack()
       
        self.errorList = {}
        self.wb = ""
        
        # Create pattern of ASCII 0-31 and trailing spaces
        whitespaces = r'[{}]'.format(''.join(chr(i) for i in range(0, 32)))
        pattern = ''.join([whitespaces, '|^ +| +$'])
        
        self.prog = re.compile(pattern)

        root.mainloop()
    
    def getExcel(self):
        self.filePath = filedialog.askopenfilename()
        self.fileLabel.set("Open " + self.filePath)
        self.wb = load_workbook(self.filePath)
        self.text.set("")

    def getNonprintable(self):
        self.errorList.clear()
        for ws in self.wb:
            lst = []    # each Excel worksheet
            for row in ws.rows:
                for cell in row:
                    s = cell.value
                    if isinstance(s, str): # only check string cells
                        if self.prog.search(s): # whitespaces found
                            lst.append(cell.coordinate)

            if lst:     # save list of cells per worksheet
                self.errorList[ws.title] = lst
            
        message = "" # display result
        for sheet in self.errorList.keys():
            message += "\n" + sheet + ":\n"
            for cell in self.errorList.get(sheet):
                message += cell + " "

        self.text.set(message)

    def removeNonprintable(self):
        if len(self.errorList) != 0: # Save if whitespace found
            for sheet in self.errorList.keys():
                ws = self.wb[sheet]
                for cell in self.errorList.get(sheet):
                    # remove whitespaces
                    s = self.prog.sub('', ws[cell].value)
                    ws[cell] = s
            
            self.filePath = self.filePath[:-5] + " edited.xlsx"
            self.wb.save(self.filePath)
            self.fileLabel.set("Remove whitespaces and Save to " + self.filePath)
            self.errorList.clear() # Clear result
        else:
            self.fileLabel.set("No whitespaces found in" + self.filePath)
Parse_Excel()
            

